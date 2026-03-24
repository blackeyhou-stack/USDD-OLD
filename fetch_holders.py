#!/usr/bin/env python3
"""
USDDOLD Token Holder Data Fetcher
=================================
Generates:
  - USDDOLD_summary_{DATE}.csv      : Cross-chain summary
  - USDDOLD_dashboard_{DATE}.xlsx   : Per-chain breakdown (Excel with merged cells)
  - {Chain}_holders_{DATE}.csv      : All fetched holders per chain

Chains covered: Tron, Ethereum, Arbitrum, Polygon, BSC (via CSV export)

Usage:
  export TRONSCAN_API_KEY="4014f55b-286c-48a7-a0eb-1dc88512b864"
  python3 fetch_holders.py
"""

import requests
import pandas as pd
from datetime import datetime
import time
import os
import binascii
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
TODAY = datetime.now().strftime('%Y%m%d')
OUTPUT_DIR = '/Users/blackey/USDDOLD/output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

TRONSCAN_API_KEY = os.environ.get('TRONSCAN_API_KEY', '')
ETHERSCAN_API_KEY = 'M29Y42G87X2W95RMAUN7797JZNS3S6R5A6'

CHAINS = {
    'Tron': {
        'contract': 'TPYmHEhy5n8TCEfYGqW2rPxsghSfzghPDn',
        'type': 'tron',
        'decimals': 18,
    },
    'Ethereum': {
        'contract': '0x0C10bF8FcB7Bf5412187A595ab97a3609160b5c6',
        'type': 'blockscout',
        'decimals': 18,
        'base_url': 'https://eth.blockscout.com',
        'verify_balances': True,   # use Etherscan to verify real balances
        'chainid': 1,
    },
    'Arbitrum': {
        'contract': '0x680447595e8b7b3Aa1B43beB9f6098C79ac2Ab3f',
        'type': 'blockscout',
        'decimals': 18,
        'base_url': 'https://arbitrum.blockscout.com',
        'verify_balances': True,
        'chainid': 42161,
    },
    'Polygon': {
        'contract': '0xffa4d863c96e743a2e1513824ea006b8d0353c57',
        'type': 'blockscout',
        'decimals': 18,
        'base_url': 'https://polygon.blockscout.com',
        'verify_balances': True,
        'chainid': 137,
    },
    'BSC': {
        'contract': '0xd17479997F34dd9156Deef8F95A52D81D265be9c',
        'type': 'bsc_csv',
        'decimals': 18,
        'csv_input_dir': '/Users/blackey/USDDOLD/input',
    },
}

# ⚙️ Boss addresses (H.E)
BOSS_ADDRESSES = {
    'Tron':     ['TT2T17KZhoDu47i2E4FWxfG79zdkEWkU9N',
                 'TPyjyZfsYaXStgz2NmAraF1uZcMtkgNan5'],
    'Arbitrum': ['0x3DdfA8eC3052539b6C9549F12cEA2C295cfF5296'],
}

REQ_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'application/json',
}

# ============================================================
# FETCH TOTAL SUPPLY (on-chain)
# ============================================================
def fetch_total_supply_tron(contract):
    """Get TRC20 totalSupply from TronScan"""
    headers = dict(REQ_HEADERS)
    if TRONSCAN_API_KEY:
        headers['TRON-PRO-API-KEY'] = TRONSCAN_API_KEY
    try:
        r = requests.get('https://apilist.tronscanapi.com/api/token_trc20',
            params={'contract': contract, 'showAll': 1},
            headers=headers, timeout=15)
        r.raise_for_status()
        items = r.json().get('trc20_tokens', [])
        if items:
            raw = float(items[0].get('total_supply_with_decimals', 0))
            decimals = int(items[0].get('decimals', 18))
            return raw / (10 ** decimals)
    except Exception as e:
        print(f"  Tron totalSupply error: {e}")
    return None


def fetch_total_supply_blockscout(base_url, contract, decimals):
    """Get ERC20 totalSupply from Blockscout token info"""
    try:
        r = requests.get(f'{base_url}/api/v2/tokens/{contract}',
            headers=REQ_HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()
        raw = data.get('total_supply', '0') or '0'
        return int(raw) / (10 ** decimals)
    except Exception as e:
        print(f"  Blockscout totalSupply error: {e}")
    return None


# ============================================================
# TRON FETCHER
# ============================================================
def fetch_tron_holders(contract, max_holders=60):
    """TronScan /api/token_trc20/holders"""
    print(f"  Fetching Tron holders...")
    holders = []
    start = 0
    page_size = 200
    total = None

    headers = dict(REQ_HEADERS)
    if TRONSCAN_API_KEY:
        headers['TRON-PRO-API-KEY'] = TRONSCAN_API_KEY
        print(f"  Using TronScan API key ✓")

    while len(holders) < max_holders:
        try:
            r = requests.get(
                'https://apilist.tronscanapi.com/api/token_trc20/holders',
                params={'contract_address': contract, 'start': start, 'limit': page_size},
                headers=headers, timeout=30,
            )
            r.raise_for_status()
            data = r.json()

            if total is None:
                total = int(data.get('total', 0))
                print(f"  Total holders reported: {total:,}")

            items = data.get('trc20_tokens', [])
            if not items:
                break

            for item in items:
                raw_bal = float(item.get('balance', 0))
                balance = raw_bal / (10 ** CHAINS['Tron']['decimals'])
                holders.append({
                    'rank': len(holders) + 1,
                    'address': item.get('holder_address', ''),
                    'name': item.get('addressTag', '') or '',
                    'balance': balance,
                    'is_contract': None,
                })

            start += len(items)
            if total and start >= total:
                break
            if len(holders) % 5000 == 0 and len(holders) > 0:
                print(f"  ... {len(holders):,} / {total:,}")
            time.sleep(0.35)

        except Exception as e:
            print(f"  Tron error at offset {start}: {e}")
            break

    print(f"  Fetched {len(holders):,} Tron holders")
    return holders


# ============================================================
# TRON ADDRESS CLASSIFICATION (TronGrid)
# ============================================================
def decode_hex_name(hex_name):
    if not hex_name:
        return ''
    try:
        return binascii.unhexlify(hex_name).decode('utf-8', errors='replace').strip('\x00')
    except Exception:
        return hex_name


def get_tron_account_info(address):
    """Check TronGrid for contract type + name, then fall back to TronScan tag1 for better labels."""
    is_contract = False
    name = ''
    try:
        r = requests.get(f'https://api.trongrid.io/v1/accounts/{address}',
            headers=REQ_HEADERS, timeout=10)
        if r.status_code == 200:
            data_list = r.json().get('data', [])
            if data_list:
                item = data_list[0]
                is_contract = item.get('type', '') == 'Contract'
                raw_name = item.get('account_name', '')
                name = decode_hex_name(raw_name) if raw_name else ''
    except Exception:
        pass

    # For contracts with generic 'CreatedByContract' name, try TronScan tag1 for proper label
    if is_contract and (not name or name == 'CreatedByContract'):
        try:
            headers = dict(REQ_HEADERS)
            if TRONSCAN_API_KEY:
                headers['TRON-PRO-API-KEY'] = TRONSCAN_API_KEY
            r2 = requests.get('https://apilist.tronscanapi.com/api/contract',
                params={'contract': address}, headers=headers, timeout=10)
            d = r2.json()
            data_list2 = d.get('data', [])
            if data_list2:
                item2 = data_list2[0]
                tag1 = item2.get('tag1', '') or item2.get('publicTag', '')
                if tag1:
                    name = tag1
        except Exception:
            pass

    return is_contract, name


# ============================================================
# BLOCKSCOUT FETCHER (ETH / ARB / Polygon)
# ============================================================
def fetch_blockscout_holders(chain_name, config, max_holders=60):
    base = config['base_url']
    contract = config['contract']
    decimals = config['decimals']
    print(f"  Fetching {chain_name} holders via Blockscout...")

    holders = []
    url = f"{base}/api/v2/tokens/{contract}/holders"
    params = {}

    while len(holders) < max_holders:
        try:
            r = requests.get(url, params=params, headers=REQ_HEADERS, timeout=30)
            r.raise_for_status()
            data = r.json()

            items = data.get('items', [])
            if not items:
                break

            for item in items:
                addr_info = item.get('address', {})
                raw_val = int(item.get('value', 0))
                balance = raw_val / (10 ** decimals)
                is_contract = bool(addr_info.get('is_contract', False))
                name = addr_info.get('name') or addr_info.get('ens_domain_name') or ''
                holders.append({
                    'rank': len(holders) + 1,
                    'address': addr_info.get('hash', ''),
                    'name': name,
                    'balance': balance,
                    'is_contract': is_contract,
                })

            next_params = data.get('next_page_params')
            if not next_params:
                break
            params = next_params
            time.sleep(0.3)

        except Exception as e:
            print(f"  {chain_name} Blockscout error: {e}")
            break

    print(f"  Fetched {len(holders):,} {chain_name} holders")
    return holders


# ============================================================
# ETHERSCAN BALANCE VERIFICATION (for ETH stale Blockscout data)
# ============================================================
def verify_eth_balances(holders, contract, chainid=1, decimals=18):
    """
    Verify each holder's real on-chain balance via Etherscan API.
    Filters out stale/incorrect Blockscout entries with zero real balance.
    Etherscan free: 5 req/sec.
    """
    print(f"  Verifying {len(holders):,} addresses via Etherscan (this takes ~{len(holders)//4} sec)...")
    verified = []
    batch_size = 5  # 5 per second max

    for i, h in enumerate(holders):
        try:
            r = requests.get('https://api.etherscan.io/v2/api', params={
                'chainid': chainid,
                'module': 'account',
                'action': 'tokenbalance',
                'contractaddress': contract,
                'address': h['address'],
                'tag': 'latest',
                'apikey': ETHERSCAN_API_KEY,
            }, timeout=10)
            data = r.json()
            if data.get('status') == '1':
                bal_raw = int(data.get('result', 0))
                real_balance = bal_raw / (10 ** decimals)
                if real_balance > 0:
                    h['balance'] = real_balance
                    verified.append(h)
            else:
                # API error — keep original if balance was significant
                if h['balance'] > 0.01:
                    verified.append(h)
        except Exception as e:
            # Network error — keep original
            if h['balance'] > 0.01:
                verified.append(h)

        if (i + 1) % 50 == 0:
            print(f"    Verified {i+1}/{len(holders)}: {len(verified)} with real balance")

        # Rate limit: 5 req/sec
        if (i + 1) % batch_size == 0:
            time.sleep(1.1)
        else:
            time.sleep(0.15)

    # Re-rank by real balance
    verified.sort(key=lambda x: x['balance'], reverse=True)
    for idx, h in enumerate(verified):
        h['rank'] = idx + 1

    print(f"  Verification done: {len(holders):,} → {len(verified):,} real holders")
    return verified


# ============================================================
# FETCH HOLDERS COUNT (for display in merged area)
# ============================================================
def fetch_holders_count_blockscout(base_url, contract):
    """Get actual total holders count from Blockscout token info"""
    try:
        r = requests.get(f'{base_url}/api/v2/tokens/{contract}',
            headers=REQ_HEADERS, timeout=15)
        r.raise_for_status()
        data = r.json()
        return int(data.get('holders_count', 0) or 0)
    except Exception:
        return 0


def fetch_holders_count_tron(contract):
    """Get actual total holders count from TronScan"""
    headers = dict(REQ_HEADERS)
    if TRONSCAN_API_KEY:
        headers['TRON-PRO-API-KEY'] = TRONSCAN_API_KEY
    try:
        r = requests.get('https://apilist.tronscanapi.com/api/token_trc20',
            params={'contract': contract, 'showAll': 1},
            headers=headers, timeout=15)
        r.raise_for_status()
        items = r.json().get('trc20_tokens', [])
        if items:
            return int(items[0].get('holders_count', 0) or 0)
    except Exception:
        pass
    return 0


# ============================================================
# BSC CSV LOADER (manual export from BSCScan)
# ============================================================
def load_bsc_from_csv(config):
    """
    Load BSC holders from a manually-exported BSCScan CSV.
    Place the CSV in the input/ directory.
    Expected columns: HolderAddress, Balance, PendingBalanceUpdate
    """
    import glob
    input_dir = config.get('csv_input_dir', '/Users/blackey/USDDOLD/input')
    os.makedirs(input_dir, exist_ok=True)

    # Find the most recent BSCScan export CSV
    pattern = os.path.join(input_dir, 'export-tokenholders-for-contract-*.csv')
    files = sorted(glob.glob(pattern))
    if not files:
        print(f"  ⚠️  No BSC CSV found in {input_dir}")
        print(f"      Download from: https://bscscan.com/token/{config['contract']}#balances")
        return []

    csv_path = files[-1]
    print(f"  Loading BSC from CSV: {os.path.basename(csv_path)}")

    try:
        df = pd.read_csv(csv_path, quotechar='"')
        # Normalize column names
        df.columns = [c.strip().strip('"') for c in df.columns]
        if 'HolderAddress' not in df.columns:
            print(f"  ⚠️  Unexpected CSV format: {list(df.columns)}")
            return []

        holders = []
        for i, row in df.iterrows():
            raw_bal = str(row.get('Balance', '0')).replace(',', '').strip()
            try:
                balance = float(raw_bal)
            except ValueError:
                continue
            if balance <= 0:
                continue
            holders.append({
                'rank': len(holders) + 1,
                'address': str(row['HolderAddress']).strip(),
                'name': '',
                'balance': balance,
                'is_contract': None,
            })

        print(f"  Loaded {len(holders):,} BSC holders from CSV")
        return holders
    except Exception as e:
        print(f"  ⚠️  BSC CSV load error: {e}")
        return []


def fetch_bsc_supply_from_page():
    """Scrape BSC USDDOLD totalSupply and holders from BSCScan page."""
    import ssl, urllib.request, re
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    contract = '0xd17479997F34dd9156Deef8F95A52D81D265be9c'
    url = f'https://bscscan.com/token/{contract}'
    req = urllib.request.Request(url, headers={
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'})
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            html = r.read().decode()
        supply_m = re.search(r'hdnTotalSupply[^>]+value="([^"]+)"', html)
        holders_m = re.search(r'Holders:\s*([\d,]+)', html)
        supply = float(supply_m.group(1).replace(',', '')) if supply_m else None
        holders = int(holders_m.group(1).replace(',', '')) if holders_m else None
        return supply, holders
    except Exception as e:
        print(f"  BSC page scrape error: {e}")
        return None, None


# ============================================================
# CLASSIFICATION ENRICHMENT
# ============================================================
def enrich_classification(chain_name, holders, config):
    """For Tron: classify top 100 via TronGrid"""
    check_limit = len(holders)
    print(f"  Classifying top {check_limit} {chain_name} addresses...")
    for h in holders[:check_limit]:
        if h.get('is_contract') is not None:
            continue
        if config['type'] == 'tron':
            is_c, name = get_tron_account_info(h['address'])
            h['is_contract'] = is_c
            if name and not h.get('name'):
                h['name'] = name
            time.sleep(0.25)
    for h in holders[check_limit:]:
        if h.get('is_contract') is None:
            h['is_contract'] = False
    return holders


# ============================================================
# ASSIGN CATEGORIES
# ============================================================
def assign_categories(chain_name, holders):
    boss_set = {a.lower() for a in BOSS_ADDRESSES.get(chain_name, [])}
    for h in holders:
        addr_lower = h['address'].lower()
        if addr_lower in boss_set:
            h['category'] = 'Boss'
        elif h.get('is_contract'):
            h['category'] = 'Protocol'
        else:
            h['category'] = 'EOA'

    eoa_rank = 0
    for h in holders:
        if h['category'] == 'EOA':
            eoa_rank += 1
            h['eoa_rank'] = eoa_rank
        else:
            h['eoa_rank'] = None
    return holders


# ============================================================
# BUILD DASHBOARD ROWS
# ============================================================
MIN_PROTOCOL_AMOUNT = 200  # contracts below this are merged

def build_dashboard(chain_name, holders, total_supply_onchain, holders_count=None, others_from_supply=False):
    """
    total_supply_onchain: fetched from contract (the real totalSupply)
    holders_count: actual total holder count from API (for display)
    """
    # Use holder-sum as denominator for percentage, but show on-chain totalSupply as column
    holder_total = sum(h['balance'] for h in holders)
    total_holders = holders_count if holders_count else len(holders)

    # Use on-chain totalSupply for percentage base (more accurate)
    pct_base = total_supply_onchain if total_supply_onchain else holder_total

    protocols = sorted([h for h in holders if h['category'] == 'Protocol'],
                       key=lambda x: x['balance'], reverse=True)
    eoas = sorted([h for h in holders if h['category'] == 'EOA'],
                  key=lambda x: x['balance'], reverse=True)
    bosses = [h for h in holders if h['category'] == 'Boss']
    top10 = eoas[:10]
    other_eoa = eoas[10:]

    rows = []

    def pct(v):
        return round(v / pct_base * 100, 4) if pct_base else 0

    ts_display = round(total_supply_onchain, 4) if total_supply_onchain else 'N/A'

    # Show individually:
    #   - Named contracts with balance >= MIN_PROTOCOL_AMOUNT
    #   - Unnamed contracts with balance >= MIN_PROTOCOL_AMOUNT (shown as address, no label)
    # Merge into 未知合约:
    #   - Named contracts with balance < MIN_PROTOCOL_AMOUNT
    #   - Unnamed contracts with balance < MIN_PROTOCOL_AMOUNT

    def is_named(h):
        return bool(h['name']) and h['name'] not in ('', 'Unknown Contract', 'CreatedByContract')

    show_individually = [h for h in protocols if h['balance'] >= MIN_PROTOCOL_AMOUNT]
    merged_contracts  = [h for h in protocols if h['balance'] < MIN_PROTOCOL_AMOUNT]

    # Sort individually-shown by balance descending
    show_individually.sort(key=lambda x: x['balance'], reverse=True)

    for h in show_individually:
        display_name = h['name'] if is_named(h) else '未知合约'
        rows.append({
            'Chain': chain_name,
            'TotalSupply': ts_display,
            'Holders': total_holders,
            'Category': 'Protocol',
            'Overall_Rank': h['rank'],
            'Address': h['address'],
            'Name/Label': display_name,
            'USDD_Amount': round(h['balance'], 4),
            'Percentage(%)': pct(h['balance']),
            'Note': '' if is_named(h) else 'unnamed contract',
        })

    # Merge small contracts (<5000 USDD)
    if merged_contracts:
        merged_total = sum(h['balance'] for h in merged_contracts)
        rows.append({
            'Chain': chain_name,
            'TotalSupply': ts_display,
            'Holders': total_holders,
            'Category': 'Protocol',
            'Overall_Rank': '',
            'Address': f'{len(merged_contracts)} contracts',
            'Name/Label': '未知合约（合并）',
            'USDD_Amount': round(merged_total, 4),
            'Percentage(%)': pct(merged_total),
            'Note': f'{len(merged_contracts)} contracts with <{MIN_PROTOCOL_AMOUNT} USDD each',
        })

    # Top 10 EOA
    for i, h in enumerate(top10, 1):
        rows.append({
            'Chain': chain_name,
            'TotalSupply': ts_display,
            'Holders': total_holders,
            'Category': 'EOA_Top10',
            'Overall_Rank': h['rank'],
            'Address': h['address'],
            'Name/Label': h['name'] or '',
            'USDD_Amount': round(h['balance'], 4),
            'Percentage(%)': pct(h['balance']),
            'Note': f'EOA #{i}',
        })

    # Other addresses (aggregated)
    if others_from_supply and total_supply_onchain:
        # "其他" = TotalSupply − sum of all top-N fetched addresses
        top_n_sum = sum(h['balance'] for h in holders)
        others_amt = max(0.0, total_supply_onchain - top_n_sum)
        others_count = (holders_count or 0) - len(holders)
        rows.append({
            'Chain': chain_name,
            'TotalSupply': ts_display,
            'Holders': total_holders,
            'Category': 'EOA_Others',
            'Overall_Rank': '',
            'Address': f'{max(0, others_count):,} addresses' if others_count > 0 else '—',
            'Name/Label': '其他地址（合并）',
            'USDD_Amount': round(others_amt, 4),
            'Percentage(%)': pct(others_amt),
            'Note': f'排名第{len(holders)+1}名及以后的所有持有者 (TotalSupply − Top{len(holders)})',
        })
    elif other_eoa:
        other_sum = sum(h['balance'] for h in other_eoa)
        rows.append({
            'Chain': chain_name,
            'TotalSupply': ts_display,
            'Holders': total_holders,
            'Category': 'EOA_Others',
            'Overall_Rank': '',
            'Address': f'{len(other_eoa)} wallets',
            'Name/Label': '其他EOA地址（合并）',
            'USDD_Amount': round(other_sum, 4),
            'Percentage(%)': pct(other_sum),
            'Note': f'EOA rank #{len(top10)+1} onwards | {len(other_eoa)} addresses',
        })

    # Boss
    for h in bosses:
        rows.append({
            'Chain': chain_name,
            'TotalSupply': ts_display,
            'Holders': total_holders,
            'Category': 'Boss',
            'Overall_Rank': h['rank'],
            'Address': h['address'],
            'Name/Label': h['name'] or 'Boss Address',
            'USDD_Amount': round(h['balance'], 4),
            'Percentage(%)': pct(h['balance']),
            'Note': 'Boss',
        })

    # Chain total footer
    rows.append({
        'Chain': chain_name,
        'TotalSupply': ts_display,
        'Holders': total_holders,
        'Category': '--- CHAIN TOTAL ---',
        'Overall_Rank': '',
        'Address': '',
        'Name/Label': 'CHAIN TOTAL',
        'USDD_Amount': round(holder_total, 4),
        'Percentage(%)': pct(holder_total),
        'Note': (f'{len(show_individually)} contracts shown | {len(merged_contracts)} merged (<{MIN_PROTOCOL_AMOUNT}) | '
                 f'{len(eoas)} EOA | {len(bosses)} boss'),
    })

    return rows, holder_total, total_holders


# ============================================================
# SAVE ALL HOLDERS CSV
# ============================================================
def save_holders(chain_name, holders, total_supply_onchain):
    """Save ALL fetched holders to CSV (renamed from save_top50)."""
    total = total_supply_onchain or sum(h['balance'] for h in holders)
    all_h = sorted(holders, key=lambda x: x['balance'], reverse=True)
    rows = []
    for i, h in enumerate(all_h, 1):
        rows.append({
            'Rank': i,
            'Address': h['address'],
            'Name/Label': h.get('name') or '',
            'USDD_Amount': round(h['balance'], 4),
            'Percentage_of_TotalSupply(%)': round(h['balance'] / total * 100, 6) if total else 0,
            'Category': h.get('category', ''),
            'Is_Contract': h.get('is_contract', ''),
        })
    df = pd.DataFrame(rows)
    fname = f"{chain_name}_holders_{TODAY}.csv"
    path = os.path.join(OUTPUT_DIR, fname)
    df.to_csv(path, index=False, encoding='utf-8-sig')
    print(f"  ✓ Saved: {fname} ({len(rows)} holders)")
    return path


# ============================================================
# EXCEL OUTPUT WITH MERGED CELLS
# ============================================================
COLS = ['Chain', 'TotalSupply', 'Holders', 'Category', 'Overall_Rank',
        'Address', 'Name/Label', 'USDD_Amount', 'Percentage(%)', 'Note']

# Color palette
COLOR_HEADER   = 'FF1F3864'  # dark navy
COLOR_PROTOCOL = 'FFDCE6F1'  # light blue
COLOR_EOA10    = 'FFE2EFDA'  # light green
COLOR_OTHERS   = 'FFFFF2CC'  # light yellow
COLOR_BOSS     = 'FFFCE4D6'  # light orange
COLOR_TOTAL    = 'FFD9D9D9'  # grey
COLOR_WHITE    = 'FFFFFFFF'

def _fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def _border():
    side = Side(style='thin', color='FFB8B8B8')
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)


def save_excel(all_dashboard_rows, filepath):
    """Generate Excel dashboard with merged cells for Chain/TotalSupply/Holders"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard'

    # ---- Header row ----
    col_widths = {
        'Chain': 12, 'TotalSupply': 22, 'Holders': 12,
        'Category': 14, 'Overall_Rank': 8,
        'Address': 44, 'Name/Label': 34,
        'USDD_Amount': 20, 'Percentage(%)': 14, 'Note': 40,
    }
    header_labels = {
        'Chain': '链', 'TotalSupply': 'TotalSupply (USDD)',
        'Holders': 'Holders', 'Category': '类型',
        'Overall_Rank': '排名', 'Address': '地址',
        'Name/Label': '名称/标签', 'USDD_Amount': 'USDD数量',
        'Percentage(%)': '占比(%)', 'Note': '备注',
    }

    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=ci, value=header_labels.get(col, col))
        cell.font = Font(bold=True, color='FFFFFFFF', size=11)
        cell.fill = _fill(COLOR_HEADER)
        cell.alignment = _center()
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col, 18)
    ws.row_dimensions[1].height = 28

    # ---- Group rows by chain for merging ----
    # Build list of (row_dict, excel_row_num)
    data_start = 2
    row_num = data_start

    # Group by chain
    chain_row_groups = {}   # chain_name -> list of row indices (1-based excel rows)
    for r in all_dashboard_rows:
        chain = r.get('Chain', '')
        if chain not in chain_row_groups:
            chain_row_groups[chain] = []
        chain_row_groups[chain].append(row_num)
        row_num += 1

    # ---- Write data rows ----
    for i, r in enumerate(all_dashboard_rows):
        er = data_start + i  # excel row
        cat = r.get('Category', '')

        if cat == '--- CHAIN TOTAL ---':
            fill = _fill(COLOR_TOTAL)
            font_kw = {'bold': True, 'size': 10}
        elif cat == 'Protocol':
            fill = _fill(COLOR_PROTOCOL)
            font_kw = {'size': 10}
        elif cat == 'EOA_Top10':
            fill = _fill(COLOR_EOA10)
            font_kw = {'size': 10}
        elif cat == 'EOA_Others':
            fill = _fill(COLOR_OTHERS)
            font_kw = {'size': 10}
        elif cat == 'Boss':
            fill = _fill(COLOR_BOSS)
            font_kw = {'bold': True, 'size': 10}
        else:
            fill = _fill(COLOR_WHITE)
            font_kw = {'size': 10}

        for ci, col in enumerate(COLS, 1):
            val = r.get(col, '')
            cell = ws.cell(row=er, column=ci, value=val)
            cell.fill = fill
            cell.font = Font(**font_kw)
            cell.border = _border()
            if col in ('Chain', 'TotalSupply', 'Holders', 'Overall_Rank', 'Percentage(%)'):
                cell.alignment = _center()
            else:
                cell.alignment = _left()

        ws.row_dimensions[er].height = 18

    # ---- Merge Chain, TotalSupply, Holders columns per chain group ----
    merge_cols = [COLS.index('Chain') + 1,
                  COLS.index('TotalSupply') + 1,
                  COLS.index('Holders') + 1]

    for chain, rows_list in chain_row_groups.items():
        if len(rows_list) <= 1:
            continue
        r_start = rows_list[0]
        r_end   = rows_list[-1]
        for mc in merge_cols:
            ws.merge_cells(
                start_row=r_start, start_column=mc,
                end_row=r_end,     end_column=mc
            )
            # Re-apply style to merged cell (top-left)
            cell = ws.cell(row=r_start, column=mc)
            cell.alignment = _center()
            cell.font = Font(bold=True, size=11)

    # ---- Freeze top row ----
    ws.freeze_panes = 'A2'

    wb.save(filepath)
    print(f"  ✓ Saved Excel: {os.path.basename(filepath)}")


# ============================================================
# MAIN
# ============================================================
def main():
    print(f"\n{'='*60}")
    print(f"  USDDOLD Token Holder Fetcher — {TODAY}")
    print(f"{'='*60}")
    print(f"  TronScan API Key: {'✓ set' if TRONSCAN_API_KEY else '✗ NOT SET'}")
    print(f"  Chains: Tron, Ethereum, Arbitrum, Polygon + BSC (CSV)")
    print(f"{'='*60}\n")

    all_dashboard_rows = []
    chain_summaries = []

    for chain_name, config in CHAINS.items():
        print(f"\n{'─'*50}")
        print(f"  Chain: {chain_name}")
        print(f"{'─'*50}")

        # 0. Fetch on-chain total supply
        if config['type'] == 'tron':
            total_supply_onchain = fetch_total_supply_tron(config['contract'])
        else:
            total_supply_onchain = fetch_total_supply_blockscout(
                config['base_url'], config['contract'], config['decimals'])
        if total_supply_onchain:
            print(f"  On-chain TotalSupply: {total_supply_onchain:,.2f} USDD")

        # 0b. Fetch actual total holders count (for display in merged area)
        if config['type'] == 'tron':
            holders_count = fetch_holders_count_tron(config['contract'])
        else:
            holders_count = fetch_holders_count_blockscout(config['base_url'], config['contract'])
        if holders_count:
            print(f"  Total Holders (API): {holders_count:,}")

        # 1. Fetch holders
        if config['type'] == 'tron':
            holders = fetch_tron_holders(config['contract'])
        else:
            holders = fetch_blockscout_holders(chain_name, config)

        if not holders:
            print(f"  ⚠️  No holder data for {chain_name}")
            chain_summaries.append({
                'Chain': chain_name,
                'Contract': config['contract'],
                'TotalSupply_Onchain': round(total_supply_onchain, 4) if total_supply_onchain else 0,
                'Holder_Sum_USDD': 0,
                'Total_Holders': holders_count or 0,
                'Percentage_of_AllChains(%)': 0,
                'Note': 'No data fetched',
            })
            continue

        # 1b. Verify real balances via Etherscan (ETH only — fixes Blockscout stale data)
        if config.get('verify_balances') and ETHERSCAN_API_KEY:
            holders = verify_eth_balances(
                holders, config['contract'],
                chainid=config.get('chainid', 1),
                decimals=config['decimals'])

        # 2. Enrich classification (Tron only)
        if config['type'] == 'tron':
            holders = enrich_classification(chain_name, holders, config)

        # 3. Assign categories
        holders = assign_categories(chain_name, holders)

        # 4. Dashboard
        dash_rows, holder_total, total_holders = build_dashboard(
            chain_name, holders, total_supply_onchain,
            holders_count=holders_count, others_from_supply=True)
        all_dashboard_rows.extend(dash_rows)

        hcount = holders_count or total_holders
        chain_summaries.append({
            'Chain': chain_name,
            'Contract': config['contract'],
            'TotalSupply_Onchain': round(total_supply_onchain, 4) if total_supply_onchain else 0,
            'Holder_Sum_USDD': round(holder_total, 4),
            'Total_Holders': hcount,
            'Percentage_of_AllChains(%)': 0,  # filled below
            'Note': '',
        })

        # 5. Save all holders CSV
        save_holders(chain_name, holders, total_supply_onchain)

        ts_str = f"{total_supply_onchain:,.2f}" if total_supply_onchain else "N/A"
        print(f"  ✓ TotalSupply: {ts_str} | "
              f"Captured: {holder_total:,.2f} | Holders: {hcount:,}")

    # ---- BSC (CSV import) ----
    print(f"\n{'─'*50}")
    print(f"  Chain: BSC (CSV import)")
    print(f"{'─'*50}")
    bsc_config = CHAINS.get('BSC', {})
    bsc_holders = load_bsc_from_csv(bsc_config)
    if bsc_holders:
        bsc_supply, bsc_holders_count = fetch_bsc_supply_from_page()
        if bsc_supply:
            print(f"  On-chain TotalSupply: {bsc_supply:,.2f} USDD")
        if bsc_holders_count:
            print(f"  Total Holders (page): {bsc_holders_count:,}")
        # BSC: mark rank1 (bridge pre-mint) as Protocol, rest classify by balance
        for h in bsc_holders:
            if h['rank'] == 1:
                h['is_contract'] = True
            else:
                h['is_contract'] = None  # unknown, classify by presence of name later
        bsc_holders = assign_categories('BSC', bsc_holders)
        dash_rows_bsc, bsc_holder_total, _ = build_dashboard(
            'BSC', bsc_holders, bsc_supply,
            holders_count=bsc_holders_count, others_from_supply=True)
        all_dashboard_rows.extend(dash_rows_bsc)
        save_holders('BSC', bsc_holders, bsc_supply)
        chain_summaries.append({
            'Chain': 'BSC',
            'Contract': bsc_config.get('contract', ''),
            'TotalSupply_Onchain': round(bsc_supply, 4) if bsc_supply else 0,
            'Holder_Sum_USDD': round(bsc_holder_total, 4),
            'Total_Holders': bsc_holders_count or 0,
            'Percentage_of_AllChains(%)': 0,
            'Note': f'Loaded from CSV',
        })
        print(f"  ✓ BSC: TotalSupply={bsc_supply:,.2f} | Holders={bsc_holders_count:,}")
    else:
        print(f"  ⚠️  BSC skipped — place CSV in {bsc_config.get('csv_input_dir', 'input/')}")

    # ---- Summary ----
    grand_total = sum(c['TotalSupply_Onchain'] for c in chain_summaries)
    for c in chain_summaries:
        c['Percentage_of_AllChains(%)'] = (
            round(c['TotalSupply_Onchain'] / grand_total * 100, 4) if grand_total else 0)

    chain_summaries.append({
        'Chain': '>>> GRAND TOTAL (5 chains)',
        'Contract': '',
        'TotalSupply_Onchain': round(grand_total, 4),
        'Holder_Sum_USDD': round(sum(c['Holder_Sum_USDD'] for c in chain_summaries
                                     if '>>>' not in c['Chain']), 4),
        'Total_Holders': sum(c['Total_Holders'] for c in chain_summaries
                              if '>>>' not in c['Chain']),
        'Percentage_of_AllChains(%)': 100.0,
        'Note': f'Data date: {TODAY} | 5 chains (BSC via CSV)',
    })

    df_summary = pd.DataFrame(chain_summaries)
    df_summary.to_csv(os.path.join(OUTPUT_DIR, f'USDDOLD_summary_{TODAY}.csv'),
                      index=False, encoding='utf-8-sig')
    print(f"\n  ✓ Saved: USDDOLD_summary_{TODAY}.csv")

    # ---- Dashboard CSV ----
    if all_dashboard_rows:
        df_dash = pd.DataFrame(all_dashboard_rows).fillna('')
        cols_csv = ['Chain', 'TotalSupply', 'Holders', 'Category', 'Overall_Rank',
                    'Address', 'Name/Label', 'USDD_Amount', 'Percentage(%)', 'Note']
        df_dash = df_dash[[c for c in cols_csv if c in df_dash.columns]]
        df_dash.to_csv(os.path.join(OUTPUT_DIR, f'USDDOLD_dashboard_{TODAY}.csv'),
                       index=False, encoding='utf-8-sig')
        print(f"  ✓ Saved: USDDOLD_dashboard_{TODAY}.csv ({len(all_dashboard_rows)} rows)")

        # ---- Dashboard XLSX (merged cells) ----
        xlsx_path = os.path.join(OUTPUT_DIR, f'USDDOLD_dashboard_{TODAY}.xlsx')
        save_excel(all_dashboard_rows, xlsx_path)

    print(f"\n{'='*60}")
    print("  SUMMARY")
    print(f"{'='*60}")
    print(df_summary.to_string(index=False))
    print(f"\n  Output: {OUTPUT_DIR}")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
